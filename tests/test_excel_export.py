import tempfile
import unittest
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

from core.models import BacklinkInfo
from db import database as db
from export.excel_export import _format_submitted_at, export_to_excel


class ExcelExportTest(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self._old_path = db.DB_PATH
        db.DB_PATH = Path(self._tmp.name) / "test.db"
        db.init_db()
        self.task_id = db.create_task(
            "Проверка example.com",
            ["example.com"],
            check_google_index=True,
            send_to_index=True,
            index_submitter="speedyindex",
        )
        db.update_task_status(self.task_id, "completed", 100)
        db.create_donors_bulk(
            self.task_id,
            [
                "https://donor-open.example/1",
                "https://donor-closed.example/2",
                "https://donor-error.example/3",
                "https://donor-skip.example/4",
            ],
        )
        donors = db.get_donors_for_task(self.task_id)
        db.update_donor(
            int(donors[0]["id"]),
            status="found",
            http_status=200,
            title="Open page",
            index_google="open",
            index_yandex="open",
            index_bing="open",
            index_baidu="closed",
            google_indexed="indexed",
            html_snippet="<html><title>Open page</title></html>",
        )
        db.update_donor(
            int(donors[1]["id"]),
            status="found",
            http_status=200,
            title="Closed page",
            index_google="closed",
            google_indexed="not_indexed",
            index_submitted_at="2026-01-01T21:00:00Z",
        )
        db.update_donor(
            int(donors[2]["id"]),
            status="found",
            http_status=200,
            title="Error page",
            index_google="open",
            google_indexed="error",
            google_index_error="XMLRiver 110: retry",
        )
        db.update_donor(
            int(donors[3]["id"]),
            status="not_loaded",
            error_code="TIMEOUT",
        )
        db.create_backlinks_bulk(
            int(donors[0]["id"]),
            self.task_id,
            [
                BacklinkInfo(
                    target_url="https://example.com/page",
                    anchor_text="купить",
                    anchor_type="text",
                    rel_type="dofollow",
                    context_html="<a>купить</a>",
                )
            ],
        )
        self._xlsx = Path(self._tmp.name) / "out.xlsx"
        export_to_excel(self.task_id, str(self._xlsx))
        self.wb = load_workbook(self._xlsx)

    def tearDown(self):
        self.wb.close()
        db.DB_PATH = self._old_path
        self._tmp.cleanup()

    def _sheet_map(self, name: str) -> list[list]:
        ws = self.wb[name]
        return [[c.value for c in row] for row in ws.iter_rows()]

    def test_workbook_has_five_named_sheets(self):
        self.assertEqual(
            self.wb.sheetnames,
            ["Сводка", "По доменам", "Доноры", "Бэклинки", "Топ анкоры"],
        )

    def test_summary_separates_robots_from_google_serp(self):
        rows = {r[0]: r[1] for r in self._sheet_map("Сводка") if r[0]}
        self.assertEqual(rows["Robots открыто (Google)"], 2)
        self.assertEqual(rows["Robots закрыто (Google)"], 1)
        self.assertEqual(rows["Robots не определено"], 1)
        self.assertEqual(rows["В индексе Google (да)"], 1)
        self.assertEqual(rows["В индексе Google (нет)"], 1)
        self.assertEqual(rows["В индексе Google (ошибка)"], 1)
        self.assertEqual(rows["В индексе Google (не проверялось)"], 1)
        self.assertEqual(rows["Отправка на индексацию"], "SpeedyIndex")
        self.assertEqual(rows["Доноров отправлено на индексацию"], 1)
        self.assertNotIn("Индекс Google", "".join(str(k) for k in rows))

    def test_donor_headers_use_robots_and_serp_labels(self):
        headers = self._sheet_map("Доноры")[0]
        self.assertIn("Robots Google", headers)
        self.assertIn("Robots Yandex", headers)
        self.assertIn("В индексе Google", headers)
        self.assertIn("Ошибка индекса Google", headers)
        self.assertIn("Отправлено на индексацию", headers)
        self.assertIn("Дата отправки на индексацию", headers)
        self.assertIn("HTML сниппет", headers)
        self.assertNotIn("Индекс Google", headers)

    def test_donor_values_are_russian_not_raw_enums(self):
        rows = self._sheet_map("Доноры")[1:]
        by_url = {r[0]: r for r in rows}
        open_row = by_url["https://donor-open.example/1"]
        closed_row = by_url["https://donor-closed.example/2"]
        error_row = by_url["https://donor-error.example/3"]
        skip_row = by_url["https://donor-skip.example/4"]

        headers = self._sheet_map("Доноры")[0]
        robots_g = headers.index("Robots Google")
        gindex = headers.index("В индексе Google")
        gerr = headers.index("Ошибка индекса Google")
        sent = headers.index("Отправлено на индексацию")
        sent_at = headers.index("Дата отправки на индексацию")

        self.assertEqual(open_row[robots_g], "Открыто")
        self.assertEqual(open_row[gindex], "Да")
        self.assertEqual(closed_row[robots_g], "Закрыто")
        self.assertEqual(closed_row[gindex], "Нет")
        self.assertEqual(closed_row[sent], "Да")
        expected_sent_at = datetime(
            2026, 1, 1, 21, 0, 0, tzinfo=timezone.utc
        ).astimezone().strftime("%d.%m.%Y %H:%M")
        self.assertEqual(closed_row[sent_at], expected_sent_at)
        self.assertEqual(open_row[sent], "Нет")
        self.assertEqual(open_row[sent_at], "—")
        self.assertEqual(skip_row[sent], "Нет")
        self.assertEqual(error_row[gindex], "Ошибка")
        self.assertEqual(error_row[gerr], "XMLRiver 110: retry")
        self.assertEqual(skip_row[robots_g], "—")
        self.assertEqual(skip_row[gindex], "—")
        self.assertNotIn("open", open_row)
        self.assertNotIn("closed", closed_row)
        self.assertNotIn("indexed", open_row)
        self.assertNotIn("not_indexed", closed_row)
        snippet_col = headers.index("HTML сниппет")
        self.assertEqual(open_row[snippet_col], "<html><title>Open page</title></html>")

    def test_domains_and_backlinks_match_in_app_report_logic(self):
        domain_rows = self._sheet_map("По доменам")
        self.assertEqual(domain_rows[0][0], "Целевой домен")
        self.assertEqual(domain_rows[1][0], "example.com")
        self.assertEqual(domain_rows[1][2], 1)
        self.assertEqual(domain_rows[1][5], "Найден")

        bl_rows = self._sheet_map("Бэклинки")
        self.assertEqual(bl_rows[1][0], "https://donor-open.example/1")
        self.assertEqual(bl_rows[1][1], "https://example.com/page")
        self.assertEqual(bl_rows[1][2], "купить")

    def test_summary_send_off_when_flag_disabled(self):
        tid = db.create_task("без отправки", ["example.com"])
        db.create_donors_bulk(tid, ["https://x.example/1"])
        path = Path(self._tmp.name) / "off.xlsx"
        export_to_excel(tid, str(path))
        wb = load_workbook(path)
        try:
            rows = {r[0]: r[1] for r in wb["Сводка"].iter_rows(values_only=True) if r[0]}
        finally:
            wb.close()
        self.assertEqual(rows["Отправка на индексацию"], "Нет")
        self.assertEqual(rows["Доноров отправлено на индексацию"], 0)

    def test_export_rejects_non_list_target_domains(self):
        with db.get_connection() as conn:
            conn.execute(
                "UPDATE tasks SET target_domains = ? WHERE id = ?",
                ('"example.com"', self.task_id),
            )
        with self.assertRaises(ValueError):
            export_to_excel(self.task_id, str(self._xlsx))


class FormatSubmittedAtTest(unittest.TestCase):
    def test_zulu_is_local_wall_time(self):
        expected = datetime(
            2026, 1, 1, 21, 0, 0, tzinfo=timezone.utc
        ).astimezone().strftime("%d.%m.%Y %H:%M")
        self.assertEqual(_format_submitted_at("2026-01-01T21:00:00Z"), expected)

    def test_offset_is_not_sliced_as_naive_utc(self):
        expected = datetime(
            2026, 1, 1, 21, 0, 0, tzinfo=timezone(timedelta(hours=5))
        ).astimezone().strftime("%d.%m.%Y %H:%M")
        self.assertEqual(
            _format_submitted_at("2026-01-01T21:00:00+05:00"), expected
        )

    def test_empty_is_dash(self):
        self.assertEqual(_format_submitted_at(""), "—")


if __name__ == "__main__":
    unittest.main()
