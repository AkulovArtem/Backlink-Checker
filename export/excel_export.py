"""
Export task results to Excel (.xlsx) with 5 sheets.
"""

import json
import logging
import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from db import database as db
from utils.url_utils import get_domain, matches_target, normalize_domain

logger = logging.getLogger(__name__)

# XML 1.0 prohibits control characters except \t \n \r; also surrogates and U+FFFE/FFFF.
# Scraped content can contain them, causing Excel's "recovery" dialog on open.
_ILLEGAL_XML_RE = re.compile("[\x00-\x08\x0b\x0c\x0e-\x1f￾￿]")


def _sanitize(val):
    """Strip illegal XML 1.0 characters from strings so Excel opens without errors."""
    if isinstance(val, str):
        return _ILLEGAL_XML_RE.sub("", val)
    return val


# Colour palette — 8-char ARGB (AARRGGBB) as required by OOXML spec.
# 6-char RGB causes Excel to show a "recovery" dialog on open.
CLR_HEADER    = "FF1E1E3A"
CLR_HEADER_FG = "FFE0E0E0"
CLR_GREEN     = "FFC8F7DC"
CLR_ORANGE    = "FFFFE0B2"
CLR_RED       = "FFFFCDD2"
CLR_ZEBRA     = "FFF5F5FF"
CLR_WHITE     = "FFFFFFFF"
CLR_BORDER    = "FFCCCCCC"


def _header_font():
    return Font(bold=True, color=CLR_HEADER_FG)


def _header_fill():
    return PatternFill(patternType="solid", fgColor=CLR_HEADER, bgColor=CLR_WHITE)


def _border():
    thin = Side(style="thin", color=CLR_BORDER)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_headers(ws, headers: list[str]):
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[1].height = 32


def _auto_width(ws, min_w=10, max_w=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:  # nosec B110
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


def _zebra_fill(row: int) -> PatternFill | None:
    return PatternFill(patternType="solid", fgColor=CLR_ZEBRA, bgColor=CLR_WHITE) if row % 2 == 0 else None


def _http_fill(status_code) -> PatternFill | None:
    if not status_code:
        return None
    grp = status_code // 100
    if grp == 2:
        return PatternFill(patternType="solid", fgColor=CLR_GREEN, bgColor=CLR_WHITE)
    if grp == 4:
        return PatternFill(patternType="solid", fgColor=CLR_ORANGE, bgColor=CLR_WHITE)
    if grp == 5:
        return PatternFill(patternType="solid", fgColor=CLR_RED, bgColor=CLR_WHITE)
    return None


def _write_row(ws, row_idx: int, values: list, zebra: bool = True):
    fill = _zebra_fill(row_idx) if zebra else None
    for col, val in enumerate(values, 1):
        sanitized = _sanitize(val)
        cell = ws.cell(row=row_idx, column=col)
        cell.value = sanitized
        if isinstance(sanitized, str):
            # openpyxl auto-sets data_type='f' for strings starting with '=';
            # override to 's' so scraped content is never treated as a formula.
            cell.data_type = "s"
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = _border()
        if fill:
            cell.fill = fill


# ── Sheet builders ─────────────────────────────────────────────────────────

_ROBOTS_RU = {"open": "Открыто", "closed": "Закрыто"}
_GINDEX_RU = {"indexed": "Да", "not_indexed": "Нет", "error": "Ошибка"}


def _donor_gindex(donor) -> str:
    try:
        return donor["google_indexed"] or ""
    except (KeyError, IndexError):
        return ""


def _donor_gindex_error(donor) -> str:
    try:
        return donor["google_index_error"] or ""
    except (KeyError, IndexError):
        return ""


def _donor_html_snippet(donor) -> str:
    try:
        return donor["html_snippet"] or ""
    except (KeyError, IndexError):
        return ""


def _robots_label(value) -> str:
    if not value:
        return "—"
    return _ROBOTS_RU.get(value, str(value))


def _gindex_label(value) -> str:
    if not value:
        return "—"
    return _GINDEX_RU.get(value, "—")


def _value_fill(value) -> PatternFill | None:
    if value in ("Открыто", "Да"):
        return PatternFill(patternType="solid", fgColor=CLR_GREEN, bgColor=CLR_WHITE)
    if value in ("Закрыто", "Нет"):
        return PatternFill(patternType="solid", fgColor=CLR_RED, bgColor=CLR_WHITE)
    if value == "Ошибка":
        return PatternFill(patternType="solid", fgColor=CLR_ORANGE, bgColor=CLR_WHITE)
    return None


def _sheet_summary(wb, task, target_domains, donors, backlinks):
    ws = wb.create_sheet("Сводка")

    total_donors = len(donors)
    df_count = sum(1 for bl in backlinks if bl["rel_type"] == "dofollow")
    nf_count  = len(backlinks) - df_count
    text_count = sum(1 for bl in backlinks if bl["anchor_type"] == "text")
    img_count  = len(backlinks) - text_count
    open_count   = sum(1 for d in donors if d["index_google"] == "open")
    closed_count = sum(1 for d in donors if d["index_google"] == "closed")
    unknown_count = total_donors - open_count - closed_count  # not loaded / pending
    g_yes = sum(1 for d in donors if _donor_gindex(d) == "indexed")
    g_no = sum(1 for d in donors if _donor_gindex(d) == "not_indexed")
    g_err = sum(1 for d in donors if _donor_gindex(d) == "error")
    g_skip = total_donors - g_yes - g_no - g_err

    _STATUS_RU = {
        "pending":   "В очереди",
        "running":   "В процессе",
        "completed": "Завершено",
        "error":     "Ошибка",
    }

    created_str = db.format_task_created(task["created_at"])

    rows = [
        ("Название задания",         task["name"]),
        ("Дата создания",            created_str),
        ("Статус",                   _STATUS_RU.get(task["status"], task["status"])),
        ("Доноров",                  total_donors),
        ("Целевые домены",           ", ".join(target_domains)),
        ("",                         ""),
        ("Всего бэклинков",          len(backlinks)),
        ("Dofollow",                 df_count),
        ("Nofollow",                 nf_count),
        ("Текстовые анкоры",         text_count),
        ("Графические анкоры",       img_count),
        ("Robots открыто (Google)",  open_count),
        ("Robots закрыто (Google)",  closed_count),
        ("Robots не определено",     unknown_count),
        ("В индексе Google (да)",    g_yes),
        ("В индексе Google (нет)",   g_no),
        ("В индексе Google (ошибка)", g_err),
        ("В индексе Google (не проверялось)", g_skip),
    ]

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 50

    for r, (key, val) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=key).font = Font(bold=True)
        sanitized = _sanitize(val)
        cell2 = ws.cell(row=r, column=2)
        cell2.value = sanitized
        if isinstance(sanitized, str):
            cell2.data_type = "s"


def _sheet_donors(wb, donors, backlinks):
    ws = wb.create_sheet("Доноры")
    headers = [
        "URL донора", "HTTP статус", "Title", "Canonical", "HTML сниппет",
        "Внутр. ссылок", "Внешн. ссылок",
        "Robots Google", "Robots Yandex", "Robots Bing", "Robots Baidu",
        "В индексе Google", "Ошибка индекса Google",
        "Найдено бэклинков"
    ]
    _write_headers(ws, headers)

    bl_counts: dict[int, int] = {}
    for bl in backlinks:
        bl_counts[bl["donor_id"]] = bl_counts.get(bl["donor_id"], 0) + 1

    for row_idx, donor in enumerate(donors, 2):
        http = donor["http_status"]
        values = [
            donor["url"],
            http or donor["error_code"] or "—",
            donor["title"] or "",
            donor["canonical_url"] or "",
            _donor_html_snippet(donor),
            donor["internal_links"] or 0,
            donor["external_links"] or 0,
            _robots_label(donor["index_google"]),
            _robots_label(donor["index_yandex"]),
            _robots_label(donor["index_bing"]),
            _robots_label(donor["index_baidu"]),
            _gindex_label(_donor_gindex(donor)),
            _donor_gindex_error(donor),
            bl_counts.get(donor["id"], 0),
        ]
        _write_row(ws, row_idx, values)
        # Apply HTTP colour to status cell
        status_cell = ws.cell(row=row_idx, column=2)
        http_fill = _http_fill(http)
        if http_fill:
            status_cell.fill = http_fill
        for col in (8, 9, 10, 11, 12):
            cell = ws.cell(row=row_idx, column=col)
            fill = _value_fill(cell.value)
            if fill:
                cell.fill = fill

    _auto_width(ws)


def _sheet_backlinks(wb, backlinks, donor_map: dict):
    ws = wb.create_sheet("Бэклинки")
    headers = [
        "URL донора", "URL цели", "Анкор",
        "Тип анкора", "Rel", "Контекст (HTML)"
    ]
    _write_headers(ws, headers)

    for row_idx, bl in enumerate(backlinks, 2):
        donor_url = donor_map.get(bl["donor_id"], "")
        _write_row(ws, row_idx, [
            donor_url,
            bl["target_url"],
            bl["anchor_text"] or "",
            bl["anchor_type"] or "",
            bl["rel_type"] or "",
            (bl["context_html"] or "")[:500],
        ])

    _auto_width(ws)


def _sheet_domains(wb, backlinks, target_domains):
    """One row per target domain: found/not-found with backlink and donor counts."""
    ws = wb.create_sheet("По доменам")
    _write_headers(ws, [
        "Целевой домен", "Доноров", "Бэклинков", "Dofollow", "Nofollow", "Статус"
    ])

    for row_idx, orig in enumerate(target_domains, 2):
        norm = normalize_domain(orig)
        matched = [
            bl for bl in backlinks
            if matches_target(bl["target_url"] or "", norm)
        ]
        donors = len({bl["donor_id"] for bl in matched})
        df = sum(1 for bl in matched if bl["rel_type"] == "dofollow")
        nf = len(matched) - df
        found = len(matched) > 0

        _write_row(ws, row_idx, [orig, donors, len(matched), df, nf,
                                  "Найден" if found else "Не найден"])
        ws.cell(row=row_idx, column=6).fill = PatternFill(
            patternType="solid",
            fgColor=CLR_GREEN if found else CLR_RED,
            bgColor=CLR_WHITE,
        )

    _auto_width(ws)


def _sheet_anchors(wb, anchor_stats):
    ws = wb.create_sheet("Топ анкоры")
    _write_headers(ws, ["Анкор", "Ссылки", "Домены", "Dofollow / Nofollow", "% от общего"])

    for row_idx, row in enumerate(anchor_stats, 2):
        _write_row(ws, row_idx, [
            row["anchor_text"] or "(пусто)",
            row["cnt"],
            row["domains"],
            f"{row['dofollow']} / {row['nofollow']}",
            f"{row['pct']:.1f}%",
        ])

    _auto_width(ws)


# ── Public API ─────────────────────────────────────────────────────────────

def export_to_excel(task_id: int, output_path: str) -> None:
    task = db.get_task(task_id)
    if not task:
        raise ValueError(f"Task {task_id} not found")

    try:
        target_domains = db.parse_target_domains(task["target_domains"])
    except (json.JSONDecodeError, TypeError) as exc:
        logger.error("Corrupted target_domains for task %d: %s", task_id, exc)
        raise ValueError(f"Task {task_id} has corrupted target_domains") from exc
    donors = db.get_donors_for_task(task_id)
    backlinks = db.get_backlinks_for_task(task_id)

    donor_map = {d["id"]: d["url"] for d in donors}

    # Rich anchor stats — same logic as report_view for consistency
    _anchor_groups: dict[str, list] = defaultdict(list)
    for _bl in backlinks:
        _anchor_groups[_bl["anchor_text"] or ""].append(_bl)
    _total_bl = len(backlinks)
    anchor_stats = []
    for _anchor, _bls in sorted(_anchor_groups.items(), key=lambda x: -len(x[1])):
        _df = sum(1 for b in _bls if b["rel_type"] == "dofollow")
        _domains = len({
            get_domain(donor_map.get(b["donor_id"], ""))
            for b in _bls
            if donor_map.get(b["donor_id"])
        })
        anchor_stats.append({
            "anchor_text": _anchor,
            "cnt":         len(_bls),
            "domains":     _domains,
            "dofollow":    _df,
            "nofollow":    len(_bls) - _df,
            "pct":         len(_bls) / _total_bl * 100 if _total_bl > 0 else 0.0,
        })

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _sheet_summary(wb, task, target_domains, donors, backlinks)
    _sheet_domains(wb, backlinks, target_domains)
    _sheet_donors(wb, donors, backlinks)
    _sheet_backlinks(wb, backlinks, donor_map)
    _sheet_anchors(wb, anchor_stats)

    wb.save(output_path)
    logger.info("Exported task %d → %s", task_id, output_path)
